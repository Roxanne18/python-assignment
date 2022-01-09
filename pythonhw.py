#The first line permits the code to import and use the "load_workbook" class
from openpyxl import load_workbook
#It is th he "openxl" library that permits enables the use and manipulation of the xlsx and csv files
wb = load_workbook('employeedata.xlsx')
#Here, in line 4 we are uploading our excel file already in the same folder with the code.
sheet = wb.active

#Here we are declaring a for loop that will iterate through the cells in rows found on the sheet
for i in range(3 ,sheet.max_row+1):
    cell = sheet.cell(i, 3)
   #here, we have the if conditions wich goes through the sheets to find the 'helpinghands.cm'
   #in any of the celss and replace it byr 'handsinhands.org'
    if 'helpinghands.cm' in cell.value:
        update = (cell.value).replace('helpinghands.cm','handsinhands.org')
        sheet.cell(i,3).value = update
#Lastly, after updating and replacing the email, the script creates a new excel sheet and saves it using the
#workbook class and save attribute    
wb.save('new_emails.xlsx')               
wb.save('new_employeedata.csv')

